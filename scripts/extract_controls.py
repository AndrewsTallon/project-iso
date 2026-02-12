#!/usr/bin/env python3
"""Extract the ISO/BSI mapping workbook into JSON for runner consumption."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any


KEY_HINTS = {
    "control_id": ["control_id", "control id", "annex", "control"],
    "title": ["title", "name", "control name"],
    "theme": ["theme", "domain", "category"],
    "automation_potential": ["automation", "potential"],
    "complexity": ["complexity"],
    "method": ["method", "approach"],
    "artifacts": ["artifact", "evidence"],
    "bsi_family": ["bsi", "grundschutz", "family"],
    "mvp_priority": ["mvp", "priority"],
}


def normalize(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def best_column_map(headers: list[str]) -> dict[str, str]:
    norm_headers = {h: normalize(h).lower() for h in headers}
    mapped: dict[str, str] = {}
    for target, hints in KEY_HINTS.items():
        for original, normalized in norm_headers.items():
            if all(hint in normalized for hint in hints[:1]) and any(hint in normalized for hint in hints):
                mapped[target] = original
                break
    return mapped


def extract_controls(workbook_path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency: openpyxl. Install it before extracting controls from Excel.") from exc

    workbook = load_workbook(workbook_path, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [normalize(cell.value) for cell in sheet[1]]
        if not any(headers):
            continue

        column_map = best_column_map(headers)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            record: dict[str, Any] = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
            record["sheet_name"] = sheet_name

            for canonical, source_col in column_map.items():
                record.setdefault(canonical, record.get(source_col))

            if not record.get("control_id"):
                continue

            rows.append(record)

    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract controls from workbook to JSON.")
    parser.add_argument("--excel", default="ISO_BSI_Compliance_Automation_Map_MVP.xlsx")
    parser.add_argument("--output", default="data/controls.json")
    args = parser.parse_args()

    try:
        controls = extract_controls(Path(args.excel))
    except RuntimeError as err:
        print(err)
        raise SystemExit(1) from err
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(controls, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(controls)} controls to {output_path}")


if __name__ == "__main__":
    main()
