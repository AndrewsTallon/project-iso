#!/usr/bin/env python3
"""Build and validate scaffold work items for multi-agent compliance execution."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency path
    load_workbook = None

SCHEMA_BY_AGENT = {
    "control_planner": "schemas/control_planner_output.schema.json",
    "collector_engineer": "schemas/collector_engineer_output.schema.json",
    "evidence_engineer": "schemas/evidence_engineer_output.schema.json",
    "security_hardening": "schemas/security_hardening_output.schema.json",
    "qa_validation": "schemas/qa_validation_output.schema.json",
    "documentation_generator": "schemas/documentation_generator_output.schema.json",
}

GLOBAL_CONSTRAINTS = {
    "deployment": "On-prem only",
    "cloud_connectors": "Prohibited for MVP",
    "hardware": "Intel N100 class, 16GB RAM, Linux appliance",
    "collection_strategy": "Agentless-first via SSH/WinRM/APIs",
    "compliance_scope": "ISO 27001 with explicit BSI IT-Grundschutz mapping",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def load_controls_index(xlsx_path: Path) -> dict[str, list[dict[str, Any]]]:
    if load_workbook is None:
        fallback = Path("data/controls.json")
        if fallback.exists():
            controls = load_json(fallback)
            index: dict[str, list[dict[str, Any]]] = {}
            for row in controls:
                control_id = str(row.get("control_id") or "").strip()
                if control_id:
                    index.setdefault(control_id, []).append(row)
            print("WARNING: openpyxl is unavailable; using data/controls.json fallback.")
            return index
        raise RuntimeError("Missing dependency: openpyxl and no data/controls.json fallback available.")

    workbook = load_workbook(xlsx_path, data_only=True)
    index: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
        if not any(header):
            continue

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            record = {header[i]: row[i] for i in range(min(len(header), len(row))) if header[i]}
            record["_sheet"] = sheet_name

            control_id = str(record.get("control_id") or record.get("Control ID") or "").strip()
            if not control_id:
                continue
            index.setdefault(control_id, []).append(record)

    return index


def load_plan(plan_path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    plan = load_json(plan_path)
    tasks = plan["task_graph"]
    modules = {
        module["module_id"]: module
        for module in plan["architecture_summary"].get("core_modules", [])
        if "module_id" in module
    }
    return tasks, modules


def parse_tasks_from_next_steps(path: Path) -> list[str]:
    if not path.exists():
        return []
    task_ids: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        match = re.search(r"\*\*(T\d+_[A-Za-z0-9_]+)", line)
        if match:
            task_ids.append(match.group(1))
    return sorted(set(task_ids))


def unblocked_tasks(tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    task_ids = {task["task_id"] for task in tasks}
    done_ids = {
        output.stem.replace(".output", "")
        for output in Path("work_outputs").glob("*.output.json")
        if output.exists()
    }
    selected: list[dict[str, Any]] = []
    for task in tasks:
        deps = task.get("blocking_dependencies", [])
        if all(dep in task_ids and dep in done_ids for dep in deps) or not deps:
            if task["task_id"] not in done_ids:
                selected.append(task)
    return selected


def build_work_item(
    task: dict[str, Any],
    modules: dict[str, dict[str, Any]],
    controls_index: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    control_scope = task.get("control_scope", [])
    controls: list[dict[str, Any]] = []
    for control_id in control_scope:
        controls.extend(controls_index.get(control_id, []))

    module = modules.get(task.get("module_id"), {"module_id": task.get("module_id"), "missing": True})
    return {
        "task": task,
        "module": module,
        "controls": controls,
        "global_constraints": GLOBAL_CONSTRAINTS,
        "expected_schema": SCHEMA_BY_AGENT.get(task["assigned_agent"]),
    }


def write_work_files(task_id: str, item: dict[str, Any], dry_run: bool = False) -> None:
    work_item_path = Path("work_items") / f"{task_id}.input.json"
    envelope_path = Path("work_outputs") / f"{task_id}.envelope.json"

    work_item_path.write_text(json.dumps(item, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    if not dry_run and not envelope_path.exists():
        placeholder = {
            "task_id": task_id,
            "run_status": "prepared",
            "errors": [],
            "generated_at": "1970-01-01T00:00:00+00:00",
            "retryable": False,
            "diagnostics": {"source": "agent_runner_placeholder"}
        }
        envelope_path.write_text(json.dumps(placeholder, indent=2) + "\n", encoding="utf-8")


def load_schema(schema_path: Path) -> dict[str, Any]:
    schema = load_json(schema_path)
    if "allOf" in schema:
        for clause in schema["allOf"]:
            ref = clause.get("$ref")
            if ref == "agent_envelope.schema.json":
                clause.clear()
                clause.update(load_json(schema_path.parent / ref))
    return schema


def validate_file(output_path: Path, agent: str | None) -> int:
    if not output_path.name.endswith(".output.json"):
        print("ERROR: --validate accepts only *.output.json contract files.")
        return 1
    try:
        from jsonschema.validators import Draft202012Validator
    except ImportError:
        print("ERROR: Missing dependency: jsonschema. Install it before using --validate.")
        return 1

    payload = load_json(output_path)
    resolved_agent = agent or payload.get("assigned_agent")
    if resolved_agent not in SCHEMA_BY_AGENT:
        print(f"ERROR: Unknown or missing assigned_agent: {resolved_agent}")
        return 1

    schema_path = Path(SCHEMA_BY_AGENT[resolved_agent])
    schema = load_schema(schema_path)
    validator = Draft202012Validator(schema)

    errors = sorted(validator.iter_errors(payload), key=lambda e: e.path)
    if errors:
        print(f"Validation FAILED for {output_path} against {schema_path}")
        for error in errors:
            path = ".".join(str(part) for part in error.absolute_path) or "<root>"
            print(f" - {path}: {error.message}")
        return 1

    print(f"Validation PASSED for {output_path} against {schema_path}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate and validate work items for agent tasks.")
    parser.add_argument("--plan", default="architecture_plan.json")
    parser.add_argument("--excel", default="ISO_BSI_Compliance_Automation_Map_MVP.xlsx")
    parser.add_argument("--next-steps", default="next_steps.md")
    parser.add_argument("--task-id", action="append", help="Specific task ID(s) to process.")
    parser.add_argument("--dry-run", action="store_true", help="Generate work items only.")
    parser.add_argument("--use-next-steps", action="store_true", help="Select tasks listed in next_steps.md.")
    parser.add_argument("--validate", help="Validate a *.output.json contract file path.")
    parser.add_argument("--agent", help="Agent type override for --validate.")
    args = parser.parse_args()

    try:
        if args.validate:
            return validate_file(Path(args.validate), args.agent)

        tasks, modules = load_plan(Path(args.plan))
        controls_index = load_controls_index(Path(args.excel))
        task_by_id = {task["task_id"]: task for task in tasks}

        if args.task_id:
            selected = [task_by_id[task_id] for task_id in args.task_id if task_id in task_by_id]
        elif args.use_next_steps:
            next_ids = parse_tasks_from_next_steps(Path(args.next_steps))
            selected = [task_by_id[task_id] for task_id in next_ids if task_id in task_by_id]
        else:
            selected = unblocked_tasks(tasks)

        if not selected:
            print("No tasks selected. Nothing to do.")
            return 0

        print(f"Selected {len(selected)} task(s).")
        for task in selected:
            item = build_work_item(task, modules, controls_index)
            write_work_files(task["task_id"], item, dry_run=args.dry_run)
            mode = "[dry-run]" if args.dry_run else "[prepared]"
            print(f"{mode} {task['task_id']} -> work_items/{task['task_id']}.input.json")

        return 0
    except FileNotFoundError as err:
        print(f"ERROR: Missing required file: {err}")
        return 1
    except Exception as err:  # pragma: no cover - defensive branch
        print(f"ERROR: Unexpected failure: {err}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
